#!/usr/bin/env python3
"""
Illinois Primary Election 2026 - Cook County Scraper
Handles Cook County Clerk (Suburban Cook County)

Platform: Custom (Excel/ZIP downloads + web interface)
Coverage: Suburban Cook County (excluding City of Chicago)

NOTE: City of Chicago has separate election authority (chicagoelections.gov)
This scraper handles ONLY the Cook County Clerk portion (suburban Cook County).

Cook County Clerk provides results in two main formats:
1. Live Excel download (election night 2026): results1124.cookcountyclerkil.gov/Home/SummaryExport
2. ZIP/Excel archive downloads (precinct canvasses page)

For March 17, 2026 Primary:
  Live Excel: https://results1124.cookcountyclerkil.gov/Home/SummaryExport
  Results site: https://results1124.cookcountyclerkil.gov/
"""

import requests
from bs4 import BeautifulSoup
import json
import re
import sys
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Run: pip install openpyxl")

class CookCountyClerkScraper:
    """Scraper for Cook County Clerk election results"""

    # March 17, 2026 Primary — live Excel export (direct download, no ZIP needed)
    LIVE_EXCEL_URL = 'https://results1124.cookcountyclerkil.gov/Home/SummaryExport'
    LIVE_RESULTS_SITE = 'https://results1124.cookcountyclerkil.gov/'
    
    def __init__(self, election_code: Optional[str] = None):
        """Initialize scraper
        
        Args:
            election_code: Legacy election code parameter (no longer used for 2026).
                          2026 Primary uses LIVE_EXCEL_URL directly.
        """
        self.election_code = election_code or 'UPDATE_ON_ELECTION_DAY'
        self.authority_name = 'Cook County Clerk'
        self.coverage = 'Suburban Cook County (excluding Chicago)'
        self.results_url = self.LIVE_RESULTS_SITE
        
        # Base URLs
        self.base_url = 'https://www.cookcountyclerkil.gov'
        self.data_url = f'{self.base_url}/elections/results-and-election-data/election-data/precinct-canvasses'

    def scrape_live_excel(self) -> Dict:
        """Scrape results from the live Excel export endpoint.
        
        For March 17, 2026 Primary:
          https://results1124.cookcountyclerkil.gov/Home/SummaryExport
        
        Returns:
            Dictionary with parsed results
        """
        if not EXCEL_AVAILABLE:
            return {
                'error': 'openpyxl not installed. Run: pip install openpyxl',
                'authority': self.authority_name,
                'scraped_at': datetime.now().isoformat()
            }

        print(f"Scraping {self.authority_name} via live Excel export...")
        print(f"URL: {self.LIVE_EXCEL_URL}")

        try:
            response = requests.get(
                self.LIVE_EXCEL_URL,
                timeout=60,
                headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
            )
            response.raise_for_status()

            workbook = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
            contests = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                contests.extend(self._parse_excel_sheet(sheet, sheet_name))

            results = {
                'authority': self.authority_name,
                'coverage': self.coverage,
                'source': f'Live Excel export — {self.LIVE_EXCEL_URL}',
                'contests': contests,
                'scraped_at': datetime.now().isoformat()
            }

            print(f"✓ Scraped {len(contests)} contests from live Excel")
            return results

        except requests.RequestException as e:
            print(f"✗ Error fetching live Excel: {e}")
            return {
                'error': str(e),
                'authority': self.authority_name,
                'recommendation': 'Try manual download: python cook_county_scraper.py --zip FILE',
                'scraped_at': datetime.now().isoformat()
            }
    
    def detect_party(self, contest_name: str) -> str:
        """Detect party affiliation from contest name
        
        Args:
            contest_name: Name of the contest
            
        Returns:
            Party string: 'Democratic', 'Republican', or 'Non-Partisan'
        """
        contest_upper = contest_name.upper()
        
        # Cook County often labels races as "DEM" or "REP" prefix
        if contest_upper.startswith('DEM ') or 'DEMOCRATIC' in contest_upper:
            return 'Democratic'
        elif contest_upper.startswith('REP ') or 'REPUBLICAN' in contest_upper:
            return 'Republican'
        else:
            return 'Non-Partisan'
    
    def parse_excel_from_zip(self, zip_path: str) -> Dict:
        """Parse election results from Excel file in ZIP archive
        
        Cook County provides precinct canvasses as ZIP files containing Excel spreadsheets.
        
        Args:
            zip_path: Path to downloaded ZIP file
            
        Returns:
            Dictionary with parsed results
        """
        results = {
            'contests': [],
            'summary': {},
            'authority': self.authority_name,
            'coverage': self.coverage,
            'source': 'Excel ZIP download',
            'scraped_at': datetime.now().isoformat()
        }
        
        print(f"Parsing Excel file from ZIP: {zip_path}")
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                # Find Excel files in ZIP
                excel_files = [f for f in zip_ref.namelist() if f.endswith(('.xlsx', '.xls'))]
                
                if not excel_files:
                    print("✗ No Excel files found in ZIP")
                    return results
                
                print(f"Found {len(excel_files)} Excel file(s)")
                
                # Process first Excel file (or could process all)
                excel_file = excel_files[0]
                print(f"Processing: {excel_file}")
                
                # Read Excel file from ZIP
                with zip_ref.open(excel_file) as f:
                    workbook = openpyxl.load_workbook(io.BytesIO(f.read()))
                    
                    # Process each sheet (often one sheet per contest or category)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        # Extract contests from this sheet
                        contests = self._parse_excel_sheet(sheet, sheet_name)
                        results['contests'].extend(contests)
                
                print(f"✓ Extracted {len(results['contests'])} contests")
                
        except Exception as e:
            print(f"✗ Error parsing ZIP file: {e}")
            results['error'] = str(e)
        
        return results
    
    def _parse_excel_sheet(self, sheet, sheet_name: str) -> List[Dict]:
        """Parse contests from an Excel sheet
        
        Cook County Excel format varies, but typically has:
        - Contest name in a header row
        - Candidate names with vote counts
        - Totals and percentages
        
        Args:
            sheet: openpyxl worksheet
            sheet_name: Name of the sheet
            
        Returns:
            List of contest dictionaries
        """
        contests = []
        current_contest = None
        
        # Iterate through rows
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Skip empty rows
            if not any(row):
                continue
            
            # Convert row to list of strings
            row_values = [str(cell) if cell is not None else '' for cell in row]
            first_cell = row_values[0].strip() if row_values else ''
            
            # Detect contest headers (typically all caps or contains specific keywords)
            if self._is_contest_header(first_cell):
                # Save previous contest
                if current_contest and current_contest.get('candidates'):
                    contests.append(current_contest)
                
                # Start new contest
                current_contest = {
                    'name': first_cell,
                    'party': self.detect_party(first_cell),
                    'sheet': sheet_name,
                    'candidates': []
                }
            
            # Detect candidate rows (have name + numbers)
            elif current_contest and len(row_values) >= 2:
                candidate_name = first_cell
                
                # Skip if looks like a header or total
                if not candidate_name or 'TOTAL' in candidate_name.upper():
                    continue
                
                # Look for vote count (numeric value in row)
                votes = 0
                percent = 0.0
                
                for cell in row_values[1:]:
                    # Try to parse as number
                    try:
                        val = float(str(cell).replace(',', '').replace('%', ''))
                        # If it has decimal, might be percentage
                        if '%' in str(cell) or (val < 100 and val > 0):
                            percent = val
                        # Otherwise it's votes
                        elif val >= 1:
                            votes = int(val)
                    except ValueError:
                        continue
                
                # Add candidate if we have votes
                if candidate_name and votes > 0:
                    current_contest['candidates'].append({
                        'name': candidate_name,
                        'votes': votes,
                        'percent': percent
                    })
        
        # Add final contest
        if current_contest and current_contest.get('candidates'):
            contests.append(current_contest)
        
        return contests
    
    def _is_contest_header(self, text: str) -> bool:
        """Determine if text is a contest header
        
        Args:
            text: Text to check
            
        Returns:
            True if this appears to be a contest name
        """
        if not text or len(text) < 5:
            return False
        
        # Contest headers are typically:
        # - All uppercase or title case
        # - Contain office names
        # - Don't start with numbers
        
        text_upper = text.upper()
        
        # Office keywords
        office_keywords = [
            'PRESIDENT', 'SENATOR', 'REPRESENTATIVE', 'GOVERNOR', 
            'CONGRESS', 'ASSEMBLY', 'HOUSE', 'SENATE', 'JUDGE',
            'BOARD', 'COMMISSIONER', 'CLERK', 'TREASURER', 'ATTORNEY',
            'DISTRICT', 'TOWNSHIP', 'COUNTY', 'STATE'
        ]
        
        # Check if contains office keyword
        has_keyword = any(keyword in text_upper for keyword in office_keywords)
        
        # Check if mostly uppercase (indicating header)
        mostly_upper = sum(1 for c in text if c.isupper()) > len(text) * 0.5
        
        return has_keyword or (mostly_upper and len(text) > 10)
    
    def scrape_web_results(self) -> Dict:
        """Scrape results from web interface
        
        Note: This may be blocked or rate-limited. Excel download is more reliable.
        
        Returns:
            Dictionary with results or error
        """
        if not self.results_url:
            return {
                'error': 'Election code not set',
                'authority': self.authority_name,
                'message': 'Set election code (e.g., "0326" for March 2026)',
                'scraped_at': datetime.now().isoformat()
            }
        
        print(f"Attempting to scrape web results from: {self.results_url}")
        print("Note: This may be blocked. Excel download is recommended.")
        
        try:
            # Try to fetch the main results page
            response = requests.get(self.results_url, timeout=30, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Parse results (format varies by election)
            # This is a placeholder - actual parsing would depend on page structure
            results = {
                'authority': self.authority_name,
                'coverage': self.coverage,
                'source': 'Web interface',
                'contests': [],
                'scraped_at': datetime.now().isoformat(),
                'note': 'Web scraping implementation varies by election - Excel download recommended'
            }
            
            return results
            
        except requests.RequestException as e:
            print(f"✗ Error fetching web results: {e}")
            return {
                'error': str(e),
                'authority': self.authority_name,
                'recommendation': 'Use Excel/ZIP download instead',
                'scraped_at': datetime.now().isoformat()
            }
    
    def save_results(self, results: Dict, output_dir: str = '.'):
        """Save results to JSON file
        
        Args:
            results: Results dictionary
            output_dir: Directory to save file
        """
        filename = f"{output_dir}/cook_clerk_results.json"
        
        with open(filename, 'w') as f:
            json.dump(results, f, indent=2)
        
        print(f"✓ Saved results to {filename}")

def print_instructions():
    """Print detailed usage instructions"""
    print("=" * 70)
    print("Cook County Clerk Election Results Scraper")
    print("Illinois Primary Election - March 17, 2026")
    print("=" * 70)
    print()
    print("COVERAGE:")
    print("  - Suburban Cook County (ALL of Cook County EXCEPT Chicago)")
    print("  - City of Chicago has separate authority: chicagoelections.gov")
    print()
    print("RECOMMENDED: Live Excel export (election night 2026)")
    print()
    print("  python cook_county_scraper.py --live")
    print()
    print("  Hits: https://results1124.cookcountyclerkil.gov/Home/SummaryExport")
    print("  Updates automatically as results come in — re-run every 15 min.")
    print()
    print("ALTERNATIVE: Manual ZIP download")
    print()
    print("  python cook_county_scraper.py --zip /path/to/downloaded.zip")
    print()

def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Cook County Clerk Election Results Scraper')
    parser.add_argument('--live', action='store_true',
                       help='Fetch live Excel export from results1124.cookcountyclerkil.gov (2026 Primary)')
    parser.add_argument('--zip', help='Path to downloaded ZIP file with Excel results')
    parser.add_argument('--code', help='Legacy: election code (e.g., 0326). Use --live instead for 2026.')
    parser.add_argument('--output', default='.', help='Output directory for JSON results')
    
    args = parser.parse_args()
    
    scraper = CookCountyClerkScraper()

    if args.live:
        results = scraper.scrape_live_excel()
        scraper.save_results(results, args.output)
    elif args.zip:
        results = scraper.parse_excel_from_zip(args.zip)
        scraper.save_results(results, args.output)
    elif args.code:
        print("⚠️  --code is no longer used for 2026. Use --live instead.")
        print("   python cook_county_scraper.py --live")
    else:
        print_instructions()

if __name__ == '__main__':
    main()
