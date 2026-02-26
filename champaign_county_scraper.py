#!/usr/bin/env python3
"""
Illinois Primary Election 2026 - Champaign County Scraper
Handles Champaign County Clerk election results

Platform: Custom web system with document downloads
Base URL: https://champaigncountyclerk.com
Coverage: Champaign County (~130K voters)

Champaign County posts results as Excel/PDF documents on their website.
Documents include: Precinct Results, County Summary, Canvass Report, etc.

Strategy: This scraper downloads and parses the "County Summary" Excel file
which contains county-wide totals for all contests.
"""

import requests
import json
import re
from datetime import datetime
from typing import Dict, List, Optional

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")

class ChampaignCountyScraper:
    """Scraper for Champaign County Clerk election results"""
    
    def __init__(self, election_date: str = '2026-03-17'):
        """Initialize scraper
        
        Args:
            election_date: Election date in YYYY-MM-DD format
        """
        self.county_name = 'Champaign'
        self.authority = 'Champaign County Clerk'
        self.election_date = election_date
        
        # Base URL - Champaign posts documents under /sites/
        self.base_url = 'https://champaigncountyclerk.com'
        self.documents_base = f'{self.base_url}/sites/champaigncountyclerk.com/files/documents'
        
        # Common document names
        self.county_summary_names = [
            'county-summary.xlsx',
            'county_summary.xlsx',
            'CountySummary.xlsx',
            'summary.xlsx'
        ]
    
    def detect_party(self, contest_name: str) -> str:
        """Detect party from contest name
        
        Champaign typically includes party in contest name or separates by ballot.
        
        Args:
            contest_name: Name of the contest
            
        Returns:
            Party string
        """
        contest_lower = contest_name.lower()
        
        # Check for explicit party markers
        if ' (dem)' in contest_lower or '(democratic' in contest_lower:
            return 'Democratic'
        elif ' (rep)' in contest_lower or '(republican' in contest_lower:
            return 'Republican'
        elif 'democratic' in contest_lower:
            return 'Democratic'
        elif 'republican' in contest_lower:
            return 'Republican'
        elif 'nonpartisan' in contest_lower or 'non-partisan' in contest_lower:
            return 'Non-Partisan'
        
        # Default to Non-Partisan for local races
        return 'Non-Partisan'
    
    def scrape_from_excel_url(self, excel_url: str) -> Dict:
        """Scrape results from County Summary Excel file
        
        Args:
            excel_url: Direct URL to Excel file
            
        Returns:
            Dictionary with scraped results
        """
        if not EXCEL_AVAILABLE:
            return {
                'error': 'openpyxl not installed',
                'message': 'Install with: pip install openpyxl',
                'authority': self.authority,
                'scraped_at': datetime.now().isoformat()
            }
        
        print(f"Scraping {self.authority}...")
        print(f"Excel URL: {excel_url}")
        
        try:
            # Download Excel file
            response = requests.get(excel_url, timeout=60)
            response.raise_for_status()
            
            # Save to temporary file and parse
            import io
            excel_file = io.BytesIO(response.content)
            
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            results = self._parse_county_summary(workbook)
            
            results['authority'] = self.authority
            results['jurisdiction'] = self.county_name
            results['election_date'] = self.election_date
            results['scraped_at'] = datetime.now().isoformat()
            results['source'] = 'County Summary Excel'
            results['excel_url'] = excel_url
            
            print(f"✓ Successfully scraped {len(results.get('contests', []))} contests")
            return results
            
        except requests.RequestException as e:
            print(f"✗ Error fetching Excel: {e}")
            return {
                'error': str(e),
                'authority': self.authority,
                'scraped_at': datetime.now().isoformat()
            }
    
    def scrape_election(self, county_summary_url: str = None) -> Dict:
        """Scrape election results
        
        Args:
            county_summary_url: Direct URL to County Summary Excel file
                               If None, returns instructions
            
        Returns:
            Dictionary with scraped results
        """
        if county_summary_url:
            return self.scrape_from_excel_url(county_summary_url)
        
        # No URL provided - return instructions
        return {
            'error': 'County Summary URL required',
            'authority': self.authority,
            'message': 'Visit https://champaigncountyclerk.com/elections/i-want-run-office/historical-election-data and find the County Summary Excel link for 2026 Primary',
            'instructions': [
                '1. Visit the Historical Election Data page',
                '2. Find "2026-03-17 - General Primary Election"',
                '3. Click "County Summary" link',
                '4. Copy the download URL',
                '5. Run: python champaign_county_scraper.py --url [URL]'
            ],
            'scraped_at': datetime.now().isoformat()
        }
    
    def _parse_county_summary(self, workbook: openpyxl.Workbook) -> Dict:
        """Parse County Summary Excel file
        
        Format varies but typically has sheets for each party/ballot type.
        Each sheet has contests with candidate names and vote totals.
        
        Args:
            workbook: openpyxl Workbook object
            
        Returns:
            Normalized results dictionary
        """
        results = {
            'contests': [],
            'summary': {}
        }
        
        # Process each sheet (typically DEM, REP, NON ballots)
        for sheet_name in workbook.sheetnames:
            print(f"  Processing sheet: {sheet_name}")
            
            sheet = workbook[sheet_name]
            party = self._detect_party_from_sheet(sheet_name)
            
            contests = self._parse_sheet_contests(sheet, party)
            results['contests'].extend(contests)
        
        return results
    
    def _detect_party_from_sheet(self, sheet_name: str) -> str:
        """Detect party from sheet name
        
        Args:
            sheet_name: Name of Excel sheet
            
        Returns:
            Party string
        """
        name_lower = sheet_name.lower()
        
        if 'dem' in name_lower or 'democratic' in name_lower:
            return 'Democratic'
        elif 'rep' in name_lower or 'republican' in name_lower:
            return 'Republican'
        elif 'non' in name_lower or 'nonpartisan' in name_lower:
            return 'Non-Partisan'
        else:
            return 'Unknown'
    
    def _parse_sheet_contests(self, sheet, default_party: str) -> List[Dict]:
        """Parse contests from an Excel sheet
        
        This is a generic parser that looks for contest headers and candidate rows.
        May need adjustment based on actual Champaign format.
        
        Args:
            sheet: openpyxl worksheet
            default_party: Party to assign if not found in contest name
            
        Returns:
            List of contest dictionaries
        """
        contests = []
        current_contest = None
        
        for row in sheet.iter_rows(values_only=True):
            if not row or not any(row):
                continue
            
            # Convert row to list, handling None values
            row = [str(cell) if cell is not None else '' for cell in row]
            
            # Check if this is a contest header (typically bold, left-aligned)
            # Usually first column has contest name
            if row[0] and not row[0].isdigit():
                # Might be a contest header
                # Check if following rows have candidate names and numbers
                
                # Save previous contest if exists
                if current_contest and current_contest.get('candidates'):
                    contests.append(current_contest)
                
                # Start new contest
                contest_name = row[0].strip()
                if contest_name and len(contest_name) > 3:
                    current_contest = {
                        'name': contest_name,
                        'party': self.detect_party(contest_name) or default_party,
                        'candidates': []
                    }
            
            # Check if this is a candidate row
            # Format: [Candidate Name] [Votes] [Percent]
            elif current_contest and len(row) >= 2:
                candidate_name = row[0].strip() if row[0] else ''
                
                # Try to find vote column (usually second or third column)
                votes = None
                for cell in row[1:4]:
                    try:
                        votes = int(str(cell).replace(',', '').replace('%', '').strip())
                        if votes > 0:
                            break
                    except (ValueError, AttributeError):
                        continue
                
                if candidate_name and votes is not None:
                    current_contest['candidates'].append({
                        'name': candidate_name,
                        'votes': votes,
                        'percent': 0  # Will calculate after
                    })
        
        # Add last contest
        if current_contest and current_contest.get('candidates'):
            contests.append(current_contest)
        
        # Calculate percentages
        for contest in contests:
            total_votes = sum(c['votes'] for c in contest['candidates'])
            if total_votes > 0:
                for candidate in contest['candidates']:
                    candidate['percent'] = round(candidate['votes'] / total_votes * 100, 2)
        
        return contests
    
    def save_results(self, results: Dict, output_dir: str = '.'):
        """Save results to JSON file
        
        Args:
            results: Results dictionary
            output_dir: Directory to save file
        """
        filename = f"{output_dir}/champaign_county_results.json"
        
        with open(filename, 'w') as f:
            json.dump(results, f, indent=2)
        
        print(f"✓ Saved results to {filename}")

def print_instructions():
    """Print detailed usage instructions"""
    print("=" * 70)
    print("Champaign County Clerk Scraper")
    print("Illinois Primary Election - March 17, 2026")
    print("=" * 70)
    print()
    print("Champaign County posts results as Excel/PDF documents.")
    print("Base URL: https://champaigncountyclerk.com")
    print()
    print("⚠️  IMPORTANT: County Summary URL Required")
    print()
    print("This scraper needs the direct URL to the County Summary Excel file.")
    print()
    print("FINDING THE URL:")
    print()
    print("1. Visit: https://champaigncountyclerk.com/elections/i-want-run-office/historical-election-data")
    print("2. Find the 2026-03-17 election date")
    print("3. Click on 'County Summary' link")
    print("4. Right-click and copy the download link")
    print("   (Should be .xlsx file)")
    print("5. Use that URL with this scraper")
    print()
    print("USAGE:")
    print()
    print("  # With County Summary URL")
    print("  python champaign_county_scraper.py --url https://champaigncountyclerk.com/sites/.../county-summary.xlsx")
    print()
    print("  # Without URL (shows these instructions)")
    print("  python champaign_county_scraper.py")
    print()
    print("DOCUMENT TYPES:")
    print()
    print("Champaign provides several document types:")
    print("- County Summary: Countywide totals (recommended)")
    print("- Precinct Results: Per-precinct detail")
    print("- Canvass Report: Official certified results")
    print("- Write-in Tally: Write-in vote details")
    print()
    print("This scraper uses the County Summary for fastest results.")
    print()

def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Champaign County Election Results Scraper')
    parser.add_argument('--url', '--county-summary-url', dest='summary_url',
                       help='Direct URL to County Summary Excel file')
    parser.add_argument('--date', default='2026-03-17',
                       help='Election date in YYYY-MM-DD format')
    parser.add_argument('--output', default='.',
                       help='Output directory for JSON results')
    
    args = parser.parse_args()
    
    scraper = ChampaignCountyScraper(args.date)
    
    if args.summary_url:
        results = scraper.scrape_from_excel_url(args.summary_url)
    else:
        print_instructions()
        results = scraper.scrape_election()
    
    scraper.save_results(results, args.output)
    
    # Print summary
    if 'error' not in results:
        print()
        print("Summary:")
        print(f"  Jurisdiction: {results.get('jurisdiction')}")
        print(f"  Authority: {results.get('authority')}")
        print(f"  Election: {results.get('election_date')}")
        print(f"  Contests: {len(results.get('contests', []))}")
    else:
        print()
        print(f"Error: {results['error']}")
        if 'instructions' in results:
            print("\nInstructions:")
            for instruction in results['instructions']:
                print(f"  {instruction}")

if __name__ == '__main__':
    main()
